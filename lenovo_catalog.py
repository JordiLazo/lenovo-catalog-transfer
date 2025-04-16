import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from tkinter import ttk
import os
from dotenv import load_dotenv
import re

# Load environment variables from .env file.
load_dotenv()
# Allowed sheet groups from .env (split by commas, trimming spaces)
allowed_sheet_names_CDEH = [s.strip() for s in os.getenv("CDEH", "").split(",") if s.strip()]
allowed_sheet_names_CDFI = [s.strip() for s in os.getenv("CDFI", "").split(",") if s.strip()]
allowed_sheet_names_CDEG = [s.strip() for s in os.getenv("CDEG", "").split(",") if s.strip()]
allowed_sheet_names_CDER = [s.strip() for s in os.getenv("CDER", "").split(",") if s.strip()]
allowed_sheet_names_CEFH = [s.strip() for s in os.getenv("CEFH", "").split(",") if s.strip()]

def log_message(message):
    """Append a log message to the log_text widget and print to console."""
    print(message)
    log_text.insert(tk.END, message + "\n")
    log_text.see(tk.END)

def get_worksheet_names(file_path: Path):
    """Reads an xlsx file and returns its list of worksheet names."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        log_message(f"Loaded source file: {file_path}")
        return sheet_names
    except Exception as e:
        log_message(f"Error reading file '{file_path}': {e}")
        messagebox.showerror("Error", f"Error reading file '{file_path}': {e}")
        return []

def open_file():
    global source_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        log_message("No source file selected.")
        return
    source_file = Path(file_path)
    log_message(f"Source file selected: {source_file}")
    sheets = get_worksheet_names(source_file)
    update_dropdown(sheets)

def select_destination():
    """Let the user pick an existing Excel file to edit."""
    global destination_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        log_message("No destination file selected.")
        return
    destination_file = Path(file_path)
    log_message(f"Destination file selected: {destination_file}")
    messagebox.showinfo("Destination Selected", f"Destination file: {destination_file}")

def update_dropdown(sheets):
    sheet_list.set(sheets)
    log_message(f"Worksheets found: {sheets}")

# Validators for each group
def is_valid_row_CDEH(cell_c, cell_d, cell_e, cell_h):
    vals = [str(cell.value or "").strip() for cell in (cell_c, cell_d, cell_e, cell_h)]
    if not all(vals):
        return False
    skip_keywords = ["part no", "family - short description", "pvpr (no iva)"]
    if any(kw in " ".join(vals).lower() for kw in skip_keywords):
        return False
    return True

def is_valid_row_CDFI(cell_c, cell_d, cell_f, cell_i):
    vals = [str(cell.value or "").strip() for cell in (cell_c, cell_d, cell_f, cell_i)]
    if not all(vals):
        return False
    skip_keywords = ["part no", "family - short description", "pvpr (no iva)"]
    if any(kw in " ".join(vals).lower() for kw in skip_keywords):
        return False
    return True

def is_valid_row_CDEG(cell_c, cell_d, cell_e, cell_g):
    vals = [str(cell.value or "").strip() for cell in (cell_c, cell_d, cell_e, cell_g)]
    if not all(vals):
        return False
    skip_keywords = ["part no", "family - short description", "pvpr (no iva)"]
    if any(kw in " ".join(vals).lower() for kw in skip_keywords):
        return False
    return True

def is_valid_row_CDER(cell_c, cell_d, cell_e, cell_r):
    vals = [str(cell.value or "").strip() for cell in (cell_c, cell_d, cell_e, cell_r)]
    if not all(vals):
        return False
    skip_keywords = ["part no", "family - short description", "pvpr (no iva)"]
    if any(kw in " ".join(vals).lower() for kw in skip_keywords):
        return False
    return True

def is_valid_row_CEFH(cell_c, cell_e, cell_f, cell_h):
    vals = [str(cell.value or "").strip() for cell in (cell_c, cell_e, cell_f, cell_h)]
    if not all(vals):
        return False
    skip_keywords = ["part no", "family - short description", "pvpr (no iva)"]
    if any(kw in " ".join(vals).lower() for kw in skip_keywords):
        return False
    return True

def get_existing_part_numbers(workbook):
    """Get all part numbers from columns C and D of the destination workbook."""
    existing_part_numbers = set()
    ws = workbook.active
    # Check all rows in columns C and D
    for row in ws.iter_rows(min_row=2, max_col=4):
        if row[2].value:  # Column C (index 2)
            existing_part_numbers.add(str(row[2].value).strip())
        if row[3].value:  # Column D (index 3)
            existing_part_numbers.add(str(row[3].value).strip())
    
    log_message(f"Found {len(existing_part_numbers)} existing part numbers in destination file")
    return existing_part_numbers

def copy_selected_sheets():
    if not source_file or not destination_file:
        messagebox.showerror("Error", "Please select both source and destination Excel files.")
        log_message("Error: Source or destination file not selected.")
        return

    selected_indices = listbox.curselection()
    selected_sheets = [listbox.get(i) for i in selected_indices]
    log_message(f"User selected sheets: {selected_sheets}")

    # Process only sheets that belong to any allowed group.
    allowed = (allowed_sheet_names_CDEH + allowed_sheet_names_CDFI +
               allowed_sheet_names_CDEG + allowed_sheet_names_CDER + allowed_sheet_names_CEFH)
    sheets_to_process = [sheet for sheet in selected_sheets if sheet in allowed]
    
    if not sheets_to_process:
        messagebox.showerror("Error", f"Please select at least one allowed sheet: {allowed}")
        log_message(f"Error: None of the selected sheets are in allowed list: {allowed}")
        return

    try:
        log_message("Starting copy operation...")
        # Open source workbook (no data_only, so we can get hyperlinks)
        src_wb = openpyxl.load_workbook(source_file, data_only=False)
        # Open destination workbook.
        try:
            dest_wb = openpyxl.load_workbook(destination_file)
            log_message(f"Opened existing destination file: {destination_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open destination file: {e}")
            log_message(f"Failed to open destination file: {e}")
            src_wb.close()
            return

        dest_ws = dest_wb.active
        
        # Get all existing part numbers from the destination file
        existing_part_numbers = get_existing_part_numbers(dest_wb)
        log_message(f"Checking for duplicates against {len(existing_part_numbers)} existing part numbers")

        # Find first empty row in destination (assuming header in row 1).
        dest_row = 2
        while dest_ws.cell(dest_row, 3).value is not None:
            dest_row += 1

        copied_rows = 0
        skipped_rows = 0
        
        for sheet_name in sheets_to_process:
            log_message(f"Processing sheet: {sheet_name}")
            src_ws = src_wb[sheet_name]
            # Process based on group membership.
            if sheet_name in allowed_sheet_names_CDEH:
                for row_cells in src_ws.iter_rows(min_row=2):
                    # CDEH uses columns: C (2), D (3), E (4), H (7)
                    cell_c = row_cells[2]
                    cell_d = row_cells[3]
                    cell_e = row_cells[4]
                    cell_h = row_cells[7]
                    if not is_valid_row_CDEH(cell_c, cell_d, cell_e, cell_h):
                        log_message("Skipping row (CDEH) due to validation failure.")
                        continue
                    part_number = str(cell_d.value or "").strip()
                    if part_number in existing_part_numbers:
                        log_message(f"Skipping duplicate Part Number (CDEH): {part_number}")
                        skipped_rows += 1
                        continue
                    val_c = str(cell_c.value or "").strip()
                    val_d = part_number
                    val_e = str(cell_e.value or "").strip()
                    val_h = cell_h.value
                    merged_val = f"{val_c} - {val_e}"
                    hyperlink = cell_d.hyperlink.target if cell_d.hyperlink else ""
                    dest_ws.cell(dest_row, 3, val_d)       # to Column C
                    dest_ws.cell(dest_row, 4, val_d)       # to Column D
                    dest_ws.cell(dest_row, 5, merged_val)  # to Column E
                    dest_ws.cell(dest_row, 8, val_h)       # to Column H
                    dest_ws.cell(dest_row, 10, hyperlink)  # to Column J
                    log_message(f"Inserted (CDEH) Part Number {val_d} at row {dest_row}")
                    existing_part_numbers.add(val_d)
                    dest_row += 1
                    copied_rows += 1
            elif sheet_name in allowed_sheet_names_CDFI:
                for row_cells in src_wb[sheet_name].iter_rows(min_row=2):
                    # CDFI uses columns: C (2), D (3), F (5), I (8)
                    cell_c = row_cells[2]
                    cell_d = row_cells[3]
                    cell_f = row_cells[5]
                    cell_i = row_cells[8]
                    if not is_valid_row_CDFI(cell_c, cell_d, cell_f, cell_i):
                        log_message("Skipping row (CDFI) due to validation failure.")
                        continue
                    part_number = str(cell_d.value or "").strip()
                    if part_number in existing_part_numbers:
                        log_message(f"Skipping duplicate Part Number (CDFI): {part_number}")
                        skipped_rows += 1
                        continue
                    val_c = str(cell_c.value or "").strip()
                    val_d = part_number
                    val_f = str(cell_f.value or "").strip()
                    val_i = cell_i.value
                    merged_val = f"{val_c} - {val_f}"
                    hyperlink = cell_d.hyperlink.target if cell_d.hyperlink else ""
                    dest_ws.cell(dest_row, 3, val_d)
                    dest_ws.cell(dest_row, 4, val_d)
                    dest_ws.cell(dest_row, 5, merged_val)
                    dest_ws.cell(dest_row, 8, val_i)
                    dest_ws.cell(dest_row, 10, hyperlink)
                    log_message(f"Inserted (CDFI) Part Number {val_d} at row {dest_row}")
                    existing_part_numbers.add(val_d)
                    dest_row += 1
                    copied_rows += 1
            elif sheet_name in allowed_sheet_names_CDEG:
                for row_cells in src_wb[sheet_name].iter_rows(min_row=2):
                    # CDEG uses columns: C (2), D (3), E (4), G (6)
                    cell_c = row_cells[2]
                    cell_d = row_cells[3]
                    cell_e = row_cells[4]
                    cell_g = row_cells[6]
                    if not is_valid_row_CDEG(cell_c, cell_d, cell_e, cell_g):
                        log_message("Skipping row (CDEG) due to validation failure.")
                        continue
                    part_number = str(cell_d.value or "").strip()
                    if part_number in existing_part_numbers:
                        log_message(f"Skipping duplicate Part Number (CDEG): {part_number}")
                        skipped_rows += 1
                        continue
                    val_c = str(cell_c.value or "").strip()
                    val_d = part_number
                    val_e = str(cell_e.value or "").strip()
                    val_g = cell_g.value
                    merged_val = f"{val_c} - {val_e}"
                    hyperlink = cell_d.hyperlink.target if cell_d.hyperlink else ""
                    dest_ws.cell(dest_row, 3, val_d)
                    dest_ws.cell(dest_row, 4, val_d)
                    dest_ws.cell(dest_row, 5, merged_val)
                    dest_ws.cell(dest_row, 8, val_g)
                    dest_ws.cell(dest_row, 10, hyperlink)
                    log_message(f"Inserted (CDEG) Part Number {val_d} at row {dest_row}")
                    existing_part_numbers.add(val_d)
                    dest_row += 1
                    copied_rows += 1
            elif sheet_name in allowed_sheet_names_CDER:
                for row_cells in src_wb[sheet_name].iter_rows(min_row=2):
                    # CDER uses columns: C (2), D (3), E (4), R (17)
                    cell_c = row_cells[2]
                    cell_d = row_cells[3]
                    cell_e = row_cells[4]
                    # Ensure the row has at least 17 cells:
                    if len(row_cells) < 17:
                        log_message("Skipping row (CDER) because it has insufficient columns.")
                        continue
                    cell_r = row_cells[16]  # Column R is index 16 (1-based: R=18)
                    if not is_valid_row_CDER(cell_c, cell_d, cell_e, cell_r):
                        log_message("Skipping row (CDER) due to validation failure.")
                        continue
                    part_number = str(cell_d.value or "").strip()
                    if part_number in existing_part_numbers:
                        log_message(f"Skipping duplicate Part Number (CDER): {part_number}")
                        skipped_rows += 1
                        continue
                    val_c = str(cell_c.value or "").strip()
                    val_d = part_number
                    val_e = str(cell_e.value or "").strip()
                    val_r = cell_r.value
                    merged_val = f"{val_c} - {val_e}"
                    hyperlink = cell_d.hyperlink.target if cell_d.hyperlink else ""
                    dest_ws.cell(dest_row, 3, val_d)
                    dest_ws.cell(dest_row, 4, val_d)
                    dest_ws.cell(dest_row, 5, merged_val)
                    dest_ws.cell(dest_row, 8, val_r)
                    dest_ws.cell(dest_row, 10, hyperlink)
                    log_message(f"Inserted (CDER) Part Number {val_d} at row {dest_row}")
                    existing_part_numbers.add(val_d)
                    dest_row += 1
                    copied_rows += 1
            elif sheet_name in allowed_sheet_names_CEFH:
                for row_cells in src_wb[sheet_name].iter_rows(min_row=2):
                    # CEFH uses columns: C (2), E (4), F (5), H (7)
                    cell_c = row_cells[2]
                    cell_e = row_cells[4]
                    cell_f = row_cells[5]
                    cell_h = row_cells[7]
                    if not is_valid_row_CEFH(cell_c, cell_e, cell_f, cell_h):
                        log_message("Skipping row (CEFH) due to validation failure.")
                        continue
                    # Here, the Part Number comes from column E (index 4) per rules.
                    part_number = str(cell_e.value or "").strip()
                    if part_number in existing_part_numbers:
                        log_message(f"Skipping duplicate Part Number (CEFH): {part_number}")
                        skipped_rows += 1
                        continue
                    val_c = str(cell_c.value or "").strip()
                    val_e = part_number  # For CEFH, column E is used as PN for destination columns C and D.
                    val_f = str(cell_f.value or "").strip()
                    val_h = cell_h.value
                    merged_val = f"{val_c} - {val_f}"
                    hyperlink = ""  # Hyperlink still from column D, but in CEFH rules it's not mentioned.
                    # If needed, you might adjust this to get hyperlink from another cell if applicable.
                    dest_ws.cell(dest_row, 3, val_e)
                    dest_ws.cell(dest_row, 4, val_e)
                    dest_ws.cell(dest_row, 5, merged_val)
                    dest_ws.cell(dest_row, 8, val_h)
                    dest_ws.cell(dest_row, 10, hyperlink)
                    log_message(f"Inserted (CEFH) Part Number {val_e} at row {dest_row}")
                    existing_part_numbers.add(val_e)
                    dest_row += 1
                    copied_rows += 1

        try:
            dest_wb.save(destination_file)
            log_message(f"Successfully copied {copied_rows} rows. Skipped {skipped_rows} duplicates.")
            messagebox.showinfo("Success", f"Data copied successfully. {copied_rows} rows copied, {skipped_rows} duplicates skipped.")
        except PermissionError as pe:
            log_message(f"Permission error: {pe}. Ensure the destination file is closed elsewhere.")
            messagebox.showerror("Error", f"Permission error: {pe}.")
        finally:
            src_wb.close()
            dest_wb.close()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to copy data: {e}")
        log_message(f"Error during copy: {e}")

def main():
    global sheet_list, listbox, log_text, source_file, destination_file
    source_file = None
    destination_file = None

    root = tk.Tk()
    root.title("Excel Sheet Selector")
    root.geometry("600x550")

    frame = tk.Frame(root)
    frame.pack(pady=10)

    tk.Button(frame, text="Select Excel File", command=open_file).pack(side=tk.LEFT, padx=5)
    tk.Button(frame, text="Select Destination Excel", command=select_destination).pack(side=tk.LEFT, padx=5)

    tk.Label(root, text="Select Worksheet(s):").pack()
    sheet_list = tk.Variable(value=[])
    listbox = tk.Listbox(root, listvariable=sheet_list, selectmode=tk.MULTIPLE, height=6)
    listbox.pack(pady=10, fill=tk.BOTH, expand=True)

    tk.Button(root, text="Copy Data", command=copy_selected_sheets).pack(pady=10)

    tk.Label(root, text="Execution Logs:").pack()
    global log_text
    log_text = tk.Text(root, height=10, state=tk.NORMAL)
    log_text.pack(pady=10, fill=tk.BOTH, expand=True)

    root.mainloop()

if __name__ == "__main__":
    main()
